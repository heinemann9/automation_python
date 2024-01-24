import openpyxl as excel
from openpyxl.styles import PatternFill
import datetime

book = excel.Workbook()
sheet = book.active

val = 3.141592
sheet.append([val,val,val])

now = datetime.datetime.now()
sheet["D1"].value = now

sheet["A1"].number_format = "0"
sheet["B1"].number_format = "0.0"
sheet["C1"].number_format = "0.00"
sheet["D1"].number_format = "yyyy년 mm월 dd일"

sheet.column_dimensions["B"].width = 100 #셀너비
sheet.row_dimensions[1].height = 100 #셀 높이

cell = sheet["A1"]
cell.fill = PatternFill(
    fill_type='solid',
    fgColor='FF0000')

book.save("python_code/autoExcel/hello.xlsx")