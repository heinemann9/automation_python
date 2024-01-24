import openpyxl as excel

book = excel.Workbook()
sheet = book.active

sheet["A1"] = "하이"

sheet.cell(row=2, column=1, value="하이2")

add_cell = sheet.cell(row=3, column=1)
add_cell.value = "하이3"

book.save("python_code/autoExcel/hello.xlsx")