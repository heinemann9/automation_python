import openpyxl as excel
import datetime

book = excel.Workbook()     #새 워크북 생성
sheet = book.active         #위 워크북에서 활성화된 시트를 가져오기

thisYear = datetime.datetime.now().year
# print(thisYear)

sheet["A1"] = "출생년도"
sheet["B1"] = "한국식 나이(old)"
sheet["C1"] = "만 나이(생일 후)"
sheet["D1"] = "만 나이(생일 전)"

sheet.column_dimensions["B"].width = 16
sheet.column_dimensions["C"].width = 16
sheet.column_dimensions["D"].width = 16

for i in range(80):
    birthYear = thisYear - i
    yearCell = sheet.cell(i + 2, 1)
    yearCell.value = str(birthYear) + "년생"
    
    koreanAge = thisYear - birthYear + 1
    ageCell = sheet.cell(i + 2, 2)
    ageCell.value = str(koreanAge) + "세"
    
    # 생일 전
    beforeAge = thisYear - birthYear - 1
    ageCell = sheet.cell(i + 2, 3)
    ageCell.value = str(beforeAge) + "세"
    
    # 생일 후
    AfterAge = thisYear - birthYear
    ageCell = sheet.cell(i + 2, 4)
    ageCell.value = str(AfterAge) + "세"

book.save("python_code/autoExcel/hello.xlsx") #저장