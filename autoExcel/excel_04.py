# import openpyxl as excel

# book = excel.Workbook()
# sheet = book.active

# for i in range(10):
#     count = i + 1
#     sheet.cell(row=count, column=1, value=count)
    
# book.save("python_code/autoExcel/hello.xlsx")

import openpyxl as excel

book = excel.Workbook()     #새 워크북 생성
sheet = book.active         #위 워크북에서 활성화된 시트를 가져오기

for i in range(2,10):
    for j in range(1,10):
        cell = sheet.cell(row=j, column=i)
        cell.value = f"{i}*{j}={i*j}"
        
        
book.save("python_code/autoExcel/hello.xlsx")