import openpyxl as excel
import datetime

book = excel.Workbook()     #새 워크북 생성
sheet = book.active         #위 워크북에서 활성화된 시트를 가져오기

thisyear = datetime.datetime.now().year #올해 연도 구하기
#print(thisyear)

#맨 윗줄에 타이틀
sheet["A1"] = "출생년도"
sheet["B1"] = "한국식 나이(old)"
sheet["C1"] = "만 나이(생일 후)"
sheet["D1"] = "만 나이(생일 전)"

#셀 너비 조정
sheet.column_dimensions["B"].width=16
sheet.column_dimensions["C"].width=16
sheet.column_dimensions["D"].width=16

#데이터 채우기
for i in range(80):
    birth_year = thisyear - i
    korean_age = thisyear - birth_year + 1
    after_day = korean_age - 1
    before_day = korean_age - 2

    year_cell = sheet.cell(i+2, 1)
    year_cell.value = str(birth_year) + "년생"
    age_cell = sheet.cell(i+2, 2)
    age_cell.value = str(korean_age) + "세"
    age2_cell = sheet.cell(i+2, 3)
    age2_cell.value = str(after_day) + "세"
    age3_cell = sheet.cell(i+2, 4)
    age3_cell.value = str(before_day) + "세"

#예외
sheet["D2"] = "-"

#파일 저장
book.save("data_excel/write_agelist.xlsx")