import openpyxl as excel

book = excel.Workbook()     #새 워크북 생성
sheet = book.active         #위 워크북에서 활성화된 시트를 가져오기

#입력 1 (하나의 셀에 입력 할 때)
sheet["A1"] = "안녕하세요. 반갑습니다."  #위치가 A1인 곳에 입력

#입력 2
sheet.cell(row=2, column=1, value="이번엔 다른 방법으로 넣었어요.")

#입력 3
add_cell = sheet.cell(row=3, column=1)
add_cell.value = "세번째 방법으로 입력 했어요."


book.save("data_excel/write_cell.xlsx") #저장