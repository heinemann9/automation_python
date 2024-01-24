import openpyxl as excel
from openpyxl.styles import PatternFill

book = excel.Workbook()     #새 워크북 생성
sheet = book.active         #위 워크북에서 활성화된 시트를 가져오기

val = 12000.141592
sheet.append([val,val,val])

sheet["A1"].number_format = "₩#,##0;[red]-₩#,##0"
sheet["B1"].number_format = "0.00"
sheet["C1"].number_format = "0.0000"
sheet["D1"].number_format = "yyyy년 mm월 dd일"
sheet["E1"].number_format = "yyyy-mm-dd"

sheet.column_dimensions["B"].width = 100 #셀너비
sheet.row_dimensions[1].height = 100 #셀 높이

cell = sheet["A1"]
cell.fill = PatternFill(
    fill_type='solid',
    fgColor='FF0000')


book.save("data_excel/write_cellformat1.xlsx") #저장